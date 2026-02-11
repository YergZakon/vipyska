"""File reader for .xlsx, .xls, and HTML-encoded .xls files."""

import os
import logging
from dataclasses import dataclass, field
from typing import List, Optional

logger = logging.getLogger('bank_parser')


@dataclass
class SheetData:
    """Represents one worksheet's data."""
    name: str
    rows: List[list] = field(default_factory=list)
    num_rows: int = 0
    num_cols: int = 0


def read_excel_file(filepath: str) -> List[SheetData]:
    """Read an Excel file, auto-detecting the actual format.

    Returns list of SheetData, one per worksheet.
    """
    ext = os.path.splitext(filepath)[1].lower()

    if ext == '.xlsx':
        return _read_xlsx(filepath)
    elif ext == '.xls':
        return _read_xls(filepath)
    else:
        raise ValueError(f"Unsupported file extension: {ext}")


def _read_xlsx(filepath: str) -> List[SheetData]:
    """Read .xlsx file using openpyxl."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"Failed to open .xlsx file {filepath}: {e}")
        # Try xlrd as fallback (file might be mislabeled)
        try:
            return _read_xls_with_xlrd(filepath)
        except Exception:
            raise e

    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))

        sd = SheetData(
            name=sheet_name,
            rows=rows,
            num_rows=len(rows),
            num_cols=max((len(r) for r in rows), default=0),
        )
        sheets.append(sd)

    wb.close()
    return sheets


def _read_xls(filepath: str) -> List[SheetData]:
    """Read .xls file — try xlrd first, then HTML fallback."""
    try:
        return _read_xls_with_xlrd(filepath)
    except Exception as e:
        logger.warning(f"xlrd failed for {filepath}: {e}, trying HTML fallback")
        try:
            return _read_xls_as_html(filepath)
        except Exception as e2:
            logger.error(f"All readers failed for {filepath}: xlrd={e}, html={e2}")
            raise


def _read_xls_with_xlrd(filepath: str) -> List[SheetData]:
    """Read .xls file using xlrd."""
    import xlrd

    wb = xlrd.open_workbook(filepath)
    sheets = []

    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        rows = []
        for row_idx in range(ws.nrows):
            row = []
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                # Convert xlrd date cells to datetime
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        dt_tuple = xlrd.xldate_as_tuple(cell.value, wb.datemode)
                        from datetime import datetime
                        row.append(datetime(*dt_tuple))
                    except Exception:
                        row.append(cell.value)
                elif cell.ctype == xlrd.XL_CELL_EMPTY:
                    row.append(None)
                else:
                    row.append(cell.value)
            rows.append(row)

        sd = SheetData(
            name=ws.name,
            rows=rows,
            num_rows=ws.nrows,
            num_cols=ws.ncols,
        )
        sheets.append(sd)

    return sheets


def _read_xls_as_html(filepath: str) -> List[SheetData]:
    """Read .xls file that is actually HTML-encoded."""
    from bs4 import BeautifulSoup

    with open(filepath, 'rb') as f:
        content = f.read()

    # Check if it looks like HTML
    if b'<html' not in content.lower() and b'<table' not in content.lower():
        raise ValueError("File is not HTML-encoded")

    soup = BeautifulSoup(content, 'lxml')
    tables = soup.find_all('table')

    if not tables:
        raise ValueError("No tables found in HTML file")

    sheets = []
    for idx, table in enumerate(tables):
        rows = []
        for tr in table.find_all('tr'):
            row = []
            for td in tr.find_all(['td', 'th']):
                text = td.get_text(strip=True)
                row.append(text if text else None)
            rows.append(row)

        sd = SheetData(
            name=f'Sheet{idx + 1}',
            rows=rows,
            num_rows=len(rows),
            num_cols=max((len(r) for r in rows), default=0),
        )
        sheets.append(sd)

    return sheets
